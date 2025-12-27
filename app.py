import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import json
import os
import io
import re
from pathlib import Path
import unicodedata

# --- 設定 ---
PAGE_TITLE = "Tシャツ＆タグ在庫管理システム"
PAGE_ICON = "👕"

# ページ設定
st.set_page_config(
    page_title=PAGE_TITLE,
    page_icon=PAGE_ICON,
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- 初期データ定義 (2025/12/14 - 12/26) ---
# ユーザー様から提供された期間のデータをプリロードします
# これにより、初回起動時からこの期間のデータが存在する状態になります
INITIAL_DATA_START = "2025-12-14"
INITIAL_DATA_END = "2025-12-26"

# 各Tシャツの定義
TSHIRT_TYPES = [
    'パンクラス×禅道会コラボTシャツ(ホワイト)ゼンプロマークなし',
    'パンクラス×禅道会コラボTシャツ(ブラック)ゼンプロマークなし',
    'パンクラス×禅道会コラボTシャツ(ホワイト)ゼンプロマークあり',
    'パンクラス×禅道会コラボTシャツ(ブラック)ゼンプロマークあり'
]

SIZES = ['150cm', '160cm', 'S', 'M', 'L', 'XL', 'XXL']

# 「黒・マークあり」の提供データ（12/14時点の在庫）
# ※他の種類は初期値0としていますが、高速インポート機能で正しいExcelを読み込めば一瞬で上書きされます
INITIAL_INVENTORY_BLACK_ARI = {
    '150cm': 10, '160cm': 5, 'S': 0, 'M': 14, 'L': 12, 'XL': 1, 'XXL': 3
}

def generate_initial_records():
    """12/14〜12/26の初期データを生成"""
    records = []
    start = datetime.strptime(INITIAL_DATA_START, "%Y-%m-%d")
    end = datetime.strptime(INITIAL_DATA_END, "%Y-%m-%d")
    
    # 日付生成
    delta = end - start
    dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(delta.days + 1)]
    # 降順（新しい日付順）にする
    dates.reverse()

    for d in dates:
        # 日付ごとの在庫データ構築
        daily_inv = {}
        for ttype in TSHIRT_TYPES:
            daily_inv[ttype] = {}
            for size in SIZES:
                # 黒・マークありの場合は初期値を設定（日が進むごとの増減はExcelインポートで補正推奨）
                if ttype == 'パンクラス×禅道会コラボTシャツ(ブラック)ゼンプロマークあり':
                    # 簡易的に12/14のデータを入れる（変動はExcel取込で修正）
                    daily_inv[ttype][size] = INITIAL_INVENTORY_BLACK_ARI.get(size, 0)
                else:
                    daily_inv[ttype][size] = 0
        
        records.append({
            'date': d,
            'timestamp': f"{d}T12:00:00",
            'inventory': daily_inv,
            'note': '初期データ'
        })
    return records

# --- カスタムCSS（視認性向上版） ---
st.markdown("""
<style>
    .main .block-container { padding-top: 2rem; padding-bottom: 5rem; }
    .stButton>button { width: 100%; border-radius: 8px; height: 3.5em; font-weight: bold; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
    .stNumberInput input { text-align: center; font-size: 1.2rem; }
    .big-number { font-size: 3rem; font-weight: bold; color: #0068c9; text-align: center; margin-bottom: 0; }
    .big-label { font-size: 1.2rem; text-align: center; opacity: 0.8; }
    div[data-testid="stExpander"] { border: 1px solid #e0e0e0; border-radius: 8px; margin-bottom: 0.8rem; }
</style>
""", unsafe_allow_html=True)

# --- 定数・パス設定 ---
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
INVENTORY_FILE = DATA_DIR / "inventory_data.json"
RECORDS_FILE = DATA_DIR / "daily_records.json"
TAG_FILE = DATA_DIR / "tag_data.json"

# --- ロジッククラス ---
class InventoryManager:
    @staticmethod
    def load_inventory():
        if INVENTORY_FILE.exists():
            try:
                with open(INVENTORY_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        # ファイルがない場合は初期データから最新の在庫を取得
        initial_records = generate_initial_records()
        if initial_records:
            return initial_records[0]['inventory'] # 最新の日付の在庫
        return {ttype: {size: 0 for size in SIZES} for ttype in TSHIRT_TYPES}
    
    @staticmethod
    def save_inventory(inventory):
        with open(INVENTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(inventory, f, ensure_ascii=False, indent=2)
    
    @staticmethod
    def load_records():
        if RECORDS_FILE.exists():
            try:
                with open(RECORDS_FILE, 'r', encoding='utf-8') as f:
                    records = json.load(f)
                    return sorted(records, key=lambda x: x['date'], reverse=True)
            except:
                pass
        
        # ファイルが存在しない場合、初期データ（12/14-12/26）を生成して返す
        print("初期データを生成します...")
        initial_data = generate_initial_records()
        # 初期データをファイルに保存しておく（永続化）
        InventoryManager.save_records(initial_data)
        return initial_data
    
    @staticmethod
    def save_records(records):
        sorted_records = sorted(records, key=lambda x: x['date'], reverse=True)
        with open(RECORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(sorted_records, f, ensure_ascii=False, indent=2)

    @staticmethod
    def load_tags():
        default_data = {"current_stock": 0, "history": []}
        if TAG_FILE.exists():
            try:
                with open(TAG_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if "history" in data:
                        data["history"] = sorted(data["history"], key=lambda x: x.get('timestamp', ''), reverse=True)
                    return data
            except:
                pass
        return default_data

    @staticmethod
    def save_tags(tag_data):
        with open(TAG_FILE, 'w', encoding='utf-8') as f:
            json.dump(tag_data, f, ensure_ascii=False, indent=2)

    @staticmethod
    def normalize_str(s):
        return unicodedata.normalize('NFC', s)

    @staticmethod
    def determine_type_from_filename(filename):
        base = InventoryManager.normalize_str(os.path.basename(filename))
        base = base.replace('（', '(').replace('）', ')')
        is_white = '白' in base or 'ホワイト' in base
        is_black = '黒' in base or 'ブラック' in base
        is_ari = 'あり' in base
        is_nasi = 'なし' in base
        
        if is_white and is_nasi: return 'パンクラス×禅道会コラボTシャツ(ホワイト)ゼンプロマークなし'
        elif is_white and is_ari: return 'パンクラス×禅道会コラボTシャツ(ホワイト)ゼンプロマークあり'
        elif is_black and is_nasi: return 'パンクラス×禅道会コラボTシャツ(ブラック)ゼンプロマークなし'
        elif is_black and is_ari: return 'パンクラス×禅道会コラボTシャツ(ブラック)ゼンプロマークあり'
        return None
    
    @staticmethod
    def normalize_size(cell_value):
        if cell_value is None: return None
        val = InventoryManager.normalize_str(str(cell_value)).strip()
        val = val.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))
        if '150' in val: return '150cm'
        if '160' in val: return '160cm'
        if 'XXL' in val or '3L' in val: return 'XXL'
        if 'XL' in val or 'LL' in val: return 'XL'
        if 'L' in val: return 'L'
        if 'M' in val: return 'M'
        if 'S' in val: return 'S'
        return None

    @staticmethod
    def import_matrix_excel_fast(uploaded_files):
        """
        【高速版】Excel/CSVインポート処理
        Pandasを使用して一括読み込みを行うため、処理が高速です。
        """
        date_records = {}
        total_loaded = 0
        
        for uploaded_file in uploaded_files:
            target_type = InventoryManager.determine_type_from_filename(uploaded_file.name)
            if not target_type: continue
            
            try:
                # ファイルタイプに応じて読み込み
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file, header=None)
                else:
                    df = pd.read_excel(uploaded_file, header=None, engine='openpyxl')
                
                # 「商品名」が含まれる行（ヘッダー行）を探す
                header_row_idx = None
                for idx, row in df.iterrows():
                    row_str = row.astype(str).values
                    if any('商品名' in s for s in row_str):
                        header_row_idx = idx
                        break
                
                if header_row_idx is None:
                    continue

                # ヘッダー行とデータ行を分割
                header = df.iloc[header_row_idx]
                data_rows = df.iloc[header_row_idx + 1:]
                
                # 日付列のマッピング作成 {col_index: 'YYYY-MM-DD'}
                date_col_map = {}
                for col_idx, val in header.items():
                    d_str = InventoryManager.parse_excel_date(val)
                    if d_str:
                        date_col_map[col_idx] = d_str
                
                if not date_col_map:
                    continue

                # データ行を反復処理
                for _, row in data_rows.iterrows():
                    # 商品名（サイズ情報）を取得（1列目か2列目にあると想定）
                    product_name_candidates = [str(row.iloc[0]), str(row.iloc[1]) if len(row) > 1 else ""]
                    product_name = next((s for s in product_name_candidates if s and s != 'nan'), "")
                    
                    size = InventoryManager.normalize_size(product_name)
                    if not size: continue
                    
                    # 日付列のデータを取得
                    for col_idx, date_str in date_col_map.items():
                        val = row.iloc[col_idx]
                        try:
                            # 文字列やNaNを0として処理
                            count = int(float(val)) if pd.notnull(val) and str(val).strip() != '' else 0
                        except:
                            count = 0
                        
                        if date_str not in date_records: date_records[date_str] = {}
                        if target_type not in date_records[date_str]: date_records[date_str][target_type] = {}
                        
                        date_records[date_str][target_type][size] = count
                        total_loaded += 1
                        
            except Exception as e:
                st.error(f"Error reading {uploaded_file.name}: {e}")
                
        return date_records, total_loaded

    @staticmethod
    def parse_excel_date(value):
        if pd.isna(value): return None
        # Excelの日付シリアル値等はPandasがdatetimeに変換している場合が多い
        if isinstance(value, datetime): return value.strftime('%Y-%m-%d')
        
        val_str = str(value).strip().replace('/', '-')
        # YYYY-MM-DD 形式チェック
        if re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', val_str):
            try: return pd.to_datetime(val_str).strftime('%Y-%m-%d')
            except: pass
        return None

# --- セッション初期化 ---
def init_session_state():
    # 読み込み順序に注意：先にRecords（初期データ含む）をロード
    if 'records' not in st.session_state:
        st.session_state.records = InventoryManager.load_records()

    if 'inventory' not in st.session_state:
        # Recordsがある場合、最新の日付のデータを現在の在庫としてセット
        if st.session_state.records:
             st.session_state.inventory = st.session_state.records[0]['inventory']
        else:
             st.session_state.inventory = InventoryManager.load_inventory()
             
    if 'tags' not in st.session_state:
        st.session_state.tags = InventoryManager.load_tags()
    if 'edit_mode' not in st.session_state:
        st.session_state.edit_mode = {}

# --- タブ1: Tシャツ在庫管理 ---
def inventory_tab():
    st.header("📦 Tシャツ在庫入力")
    today = datetime.now().strftime("%Y-%m-%d")
    
    last_record_date = st.session_state.records[0]['date'] if st.session_state.records else "なし"
    if last_record_date != today:
        st.warning(f"⚠️ 本日 ({today}) の記録がまだ保存されていません。（最終記録: {last_record_date}）")
    else:
        st.success(f"✅ 本日 ({today}) の記録は保存済みです。")

    col_act1, col_act2 = st.columns(2)
    with col_act1:
        if st.button("💾 本日の記録を保存/更新", type="primary", use_container_width=True):
            save_daily_record()
    with col_act2:
        if st.button("📤 Tシャツ在庫をExcelでDL", use_container_width=True):
            export_current_excel()

    st.markdown("---")
    # 高速化されたインポート機能
    with st.expander("📥 過去データをExcel/CSVから一括インポート（高速版）"):
        st.info("※ 処理を最適化しました。大量のデータも数秒で反映されます。")
        uploaded_files = st.file_uploader("ファイルをドラッグ&ドロップ", type=['xlsx', 'xls', 'csv'], accept_multiple_files=True)
        if uploaded_files:
            import_excel_data(uploaded_files)

    st.markdown("### 在庫数入力")
    for ttype in TSHIRT_TYPES:
        display_name = ttype.replace('パンクラス×禅道会コラボTシャツ', '').replace('ゼンプロマーク', 'マーク')
        with st.container():
            st.markdown(f"**{display_name}**")
            cols = st.columns(len(SIZES))
            for idx, size in enumerate(SIZES):
                with cols[idx]:
                    current_val = st.session_state.inventory.get(ttype, {}).get(size, 0)
                    new_val = st.number_input(f"{size}", min_value=0, value=current_val, step=1, key=f"inv_{ttype}_{size}")
                    if new_val != current_val:
                        st.session_state.inventory[ttype][size] = new_val
                        InventoryManager.save_inventory(st.session_state.inventory)
                    
                    c_minus, c_plus = st.columns(2)
                    if c_minus.button("－", key=f"m_{ttype}_{size}"):
                        st.session_state.inventory[ttype][size] = max(0, current_val - 1)
                        InventoryManager.save_inventory(st.session_state.inventory)
                        st.rerun()
                    if c_plus.button("＋", key=f"p_{ttype}_{size}"):
                        st.session_state.inventory[ttype][size] = current_val + 1
                        InventoryManager.save_inventory(st.session_state.inventory)
                        st.rerun()
            st.markdown("---")

def save_daily_record():
    today = datetime.now().strftime("%Y-%m-%d")
    existing_idx = None
    for idx, record in enumerate(st.session_state.records):
        if record['date'] == today:
            existing_idx = idx
            break
    new_record = {
        'date': today,
        'timestamp': datetime.now().isoformat(),
        'inventory': json.loads(json.dumps(st.session_state.inventory)),
        'note': '手動保存'
    }
    if existing_idx is not None:
        st.session_state.records[existing_idx] = new_record
        st.toast(f"✅ {today}の記録を更新しました")
    else:
        st.session_state.records.insert(0, new_record)
        st.toast(f"✅ {today}の記録を新規保存しました")
    InventoryManager.save_records(st.session_state.records)
    st.rerun()

def import_excel_data(uploaded_files):
    # 高速版メソッドを使用
    date_records, total_loaded = InventoryManager.import_matrix_excel_fast(uploaded_files)
    
    if date_records:
        existing_map = {r['date']: r for r in st.session_state.records}
        for date_str, type_data in date_records.items():
            if date_str in existing_map:
                record = existing_map[date_str]
                for ttype, sizes in type_data.items():
                    if ttype not in record['inventory']: record['inventory'][ttype] = {s: 0 for s in SIZES}
                    for s, count in sizes.items(): record['inventory'][ttype][s] = count
            else:
                new_inventory = {t: {s: 0 for s in SIZES} for t in TSHIRT_TYPES}
                for ttype, sizes in type_data.items():
                    for s, count in sizes.items(): new_inventory[ttype][s] = count
                new_record = {'date': date_str, 'timestamp': f"{date_str}T12:00:00", 'inventory': new_inventory, 'note': 'Excel自動取込'}
                st.session_state.records.append(new_record)
        
        # データを日付順にソートし直す
        st.session_state.records.sort(key=lambda x: x['date'], reverse=True)
        InventoryManager.save_records(st.session_state.records)
        
        # もし最新日付のデータが更新されていたら、現在の在庫表示にも反映
        if st.session_state.records:
             st.session_state.inventory = st.session_state.records[0]['inventory']
             InventoryManager.save_inventory(st.session_state.inventory)
             
        st.success(f"✅ インポート完了: {len(date_records)}日分のデータを高速処理しました。（更新セル数: {total_loaded}）")
        st.rerun()
    else:
        st.error("⚠️ データが見つかりませんでした。ファイル形式を確認してください。")

def export_current_excel():
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    for i, ttype in enumerate(TSHIRT_TYPES):
        safe_title = ttype[:30].replace('/', '_')
        if i == 0: ws = wb.active; ws.title = safe_title
        else: ws = wb.create_sheet(title=safe_title)
        ws.append(['サイズ', '在庫数'])
        for j, size in enumerate(SIZES):
            ws.cell(row=j+2, column=1, value=size)
            ws.cell(row=j+2, column=2, value=st.session_state.inventory[ttype].get(size, 0))
    wb.save(output)
    output.seek(0)
    st.download_button("📥 Excelダウンロード", output, f"在庫_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- タブ2: タグ管理 ---
def tags_tab():
    st.header("🏷️ タグ（衣服）在庫管理")
    
    current_stock = st.session_state.tags.get("current_stock", 0)
    
    st.markdown("<div class='big-label'>現在の在庫数</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='big-number'>{current_stock:,} 枚</div>", unsafe_allow_html=True)
    st.markdown("---")

    st.subheader("📝 在庫の更新（使用・入荷・不良）")
    st.caption("※ タグを使用した日、または入荷した際にここから入力してください。")

    with st.form("tag_action_form", clear_on_submit=True):
        col1, col2 = st.columns([1, 2])
        with col1:
            action_type = st.radio("区分", ["使用 (－)", "入荷・追加 (＋)", "不良 (－)"], horizontal=False)
        with col2:
            amount = st.number_input("数量 (枚)", min_value=1, step=1, value=1)
            note = st.text_input("備考 (任意)", placeholder="例: 12月分受注, 追加発注分など")
        
        submitted = st.form_submit_button("更新を記録する", use_container_width=True)
        
        if submitted and amount > 0:
            update_tag_stock(action_type, amount, note)
    
    st.markdown("---")
    
    st.subheader("📜 更新履歴")
    history = st.session_state.tags.get("history", [])
    if history:
        df_hist = pd.DataFrame(history)
        st.dataframe(df_hist, use_container_width=True)
    else:
        st.info("まだ履歴がありません。")

def update_tag_stock(action_type, amount, note):
    current_stock = st.session_state.tags.get("current_stock", 0)
    
    if "使用" in action_type:
        new_stock = current_stock - amount
        act_label = "使用"
    elif "入荷" in action_type:
        new_stock = current_stock + amount
        act_label = "入荷"
    elif "不良" in action_type:
        new_stock = current_stock - amount
        act_label = "不良"
    
    if new_stock < 0:
        st.warning("⚠️ 在庫数がマイナスになります。")

    new_entry = {
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "date": datetime.now().strftime('%Y-%m-%d'),
        "action": act_label,
        "amount": amount,
        "stock_after": new_stock,
        "note": note
    }
    
    st.session_state.tags["current_stock"] = new_stock
    st.session_state.tags["history"].insert(0, new_entry)
    
    InventoryManager.save_tags(st.session_state.tags)
    st.success(f"✅ {act_label} {amount}枚 を記録しました。（現在庫: {new_stock}枚）")
    st.rerun()

# --- タブ3: Tシャツ日次記録 ---
def records_tab():
    st.header("📊 Tシャツ日次記録")
    with st.expander("🔎 期間で絞り込み", expanded=False):
        c1, c2 = st.columns(2)
        start_date = c1.date_input("開始", value=datetime.now() - timedelta(days=60))
        end_date = c2.date_input("終了", value=datetime.now())
    
    c_csv, c_xls = st.columns(2)
    with c_csv:
        if st.button("📊 CSVダウンロード", use_container_width=True): export_records('csv', start_date, end_date)
    with c_xls:
        if st.button("📈 Excelダウンロード", use_container_width=True): export_records('excel', start_date, end_date)
    st.divider()

    records = st.session_state.records
    if not records:
        st.info("データがありません。")
        return

    for i, record in enumerate(records):
        d_str = record['date']
        if not (start_date <= datetime.strptime(d_str, '%Y-%m-%d').date() <= end_date): continue
        note = record.get('note', '')
        with st.expander(f"📅 {d_str} {f'({note})' if note else ''}"):
            is_editing = st.session_state.edit_mode.get(d_str, False)
            c_info, c_edit, c_del = st.columns([6, 2, 2])
            with c_edit:
                if st.button("✏️ 編集", key=f"btn_edit_{d_str}"):
                    st.session_state.edit_mode[d_str] = not is_editing
                    st.rerun()
            with c_del:
                if st.button("🗑️ 削除", key=f"btn_del_{d_str}", type="primary"):
                    st.session_state.records.pop(i)
                    InventoryManager.save_records(st.session_state.records)
                    st.rerun()
            
            if is_editing:
                st.info("📝 編集中...")
                edited_inv = record['inventory'].copy()
                for ttype in TSHIRT_TYPES:
                    st.caption(f"**{ttype}**")
                    cols = st.columns(len(SIZES))
                    for idx, size in enumerate(SIZES):
                        key = f"e_{d_str}_{ttype}_{size}"
                        old_val = edited_inv.get(ttype, {}).get(size, 0)
                        edited_inv[ttype][size] = cols[idx].number_input(size, value=old_val, min_value=0, key=key, label_visibility="collapsed")
                if st.button("💾 保存", key=f"save_{d_str}"):
                    record['inventory'] = edited_inv
                    InventoryManager.save_records(st.session_state.records)
                    st.session_state.edit_mode[d_str] = False
                    st.rerun()
            else:
                st.dataframe(pd.DataFrame([{"種類": t.replace('パンクラス×禅道会コラボTシャツ', ''), **inv} for t, inv in record['inventory'].items()]).set_index("種類"))

def export_records(fmt, start, end):
    records = st.session_state.records
    data = []
    for r in records:
        d = r['date']
        if not (start <= datetime.strptime(d, '%Y-%m-%d').date() <= end): continue
        for ttype in TSHIRT_TYPES:
            for size in SIZES:
                data.append({"日付": d, "種類": ttype, "サイズ": size, "在庫数": r['inventory'].get(ttype, {}).get(size, 0)})
    df = pd.DataFrame(data)
    if df.empty:
        st.warning("対象データなし")
        return
    if fmt == 'csv':
        st.download_button("CSV DL", df.to_csv(index=False).encode('utf-8-sig'), "records.csv", "text/csv")
    else:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.pivot_table(index=['種類', 'サイズ'], columns='日付', values='在庫数', fill_value=0).to_excel(writer, sheet_name="日次推移")
        output.seek(0)
        st.download_button("Excel DL", output, "records.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- タブ4: データ管理 ---
def settings_tab():
    st.header("⚙️ データ管理")
    st.warning("クラウド版（Web）では再起動でデータが消えるため、定期的にバックアップをDLしてください。")
    
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📤 バックアップ")
        full_data = {
            'inventory': st.session_state.inventory,
            'records': st.session_state.records,
            'tags': st.session_state.tags,
            'saved_at': datetime.now().isoformat()
        }
        json_str = json.dumps(full_data, ensure_ascii=False, indent=2)
        st.download_button("📦 全データをバックアップ", json_str, f"backup_{datetime.now().strftime('%Y%m%d')}.json", "application/json", type="primary")

    with col2:
        st.subheader("📥 データ復元")
        uploaded = st.file_uploader("バックアップファイル (.json)", type=['json'])
        if uploaded:
            try:
                data = json.load(uploaded)
                if 'inventory' in data: st.session_state.inventory = data['inventory']
                if 'records' in data: st.session_state.records = data['records']
                if 'tags' in data: st.session_state.tags = data['tags']
                
                InventoryManager.save_inventory(st.session_state.inventory)
                InventoryManager.save_records(st.session_state.records)
                InventoryManager.save_tags(st.session_state.tags)
                
                st.success("✅ データを復元しました！")
                if st.button("更新を反映"): st.rerun()
            except Exception as e:
                st.error(f"復元失敗: {e}")

# --- タブ5: マニュアル ---
def manual_tab():
    st.header("📖 システム操作マニュアル")
    st.markdown("""
    このシステムは、**「Tシャツ」**と**「タグ」**の在庫を管理し、記録を残すためのツールです。
    データが消えないよう、以下の手順に従って操作してください。
    """)

    with st.expander("1. Tシャツの在庫管理（毎日実施）", expanded=True):
        st.markdown("""
        **【概要】**
        * 毎日、その時点でのTシャツ在庫数を入力し、保存します。
        
        **【手順】**
        1.  **「📦 Tシャツ在庫」**タブを開きます。
        2.  各Tシャツのサイズごとに、現在の在庫数を入力します（＋－ボタンも使えます）。
        3.  入力が終わったら、画面上部の**「💾 本日の記録を保存/更新」**ボタンを押します。
        4.  画面右上に「✅ 保存しました」と表示されれば完了です。
        
        **【高速インポート】**
        * 今回のアップデートで、Excel/CSVの読み込みが**劇的に高速化**しました。
        * 「📥 過去データをExcel/CSVから一括インポート」にファイルをドラッグすると、数秒で反映されます。
        """)

    with st.expander("2. タグ（衣服）の在庫管理（使用・入荷時のみ）", expanded=True):
        st.markdown("""
        **【概要】**
        * タグを使用した日や、新しいタグが入荷した時に記録します。
        * 日々の入力は不要です。アクションがあった時だけ操作してください。
        
        **【手順】**
        1.  **「🏷️ タグ管理」**タブを開きます。
        2.  フォームで**「使用」「入荷」「不良」**のいずれかを選択します。
        3.  枚数を入力し、必要であれば備考（「〇月分受注」など）を記入します。
        4.  **「更新を記録する」**ボタンを押します。
        """)

    with st.expander("3. データの修正・確認", expanded=True):
        st.markdown("""
        * **Tシャツの履歴:** 「📊 Tシャツ記録」タブで過去の記録を確認できます。「✏️ 編集」ボタンで後から数値を修正したり、「🗑️ 削除」で間違った日の記録を消すことができます。
        * **データの出力:** 各タブにある「Excelダウンロード」等のボタンから、報告用のファイルを作成できます。
        """)

    with st.expander("4. 【重要】データのバックアップと復元", expanded=True):
        st.warning("⚠️ この作業は非常に重要です")
        st.markdown("""
        このシステムはWeb上で動作しているため、**長時間放置したりページを閉じたりすると、入力したデータがリセットされる場合があります。**
        
        **【作業終了時】**
        1.  **「⚙️ データ管理」**タブを開きます。
        2.  **「📦 全データをバックアップ」**ボタンを押し、ファイルをPCやiPadに保存してください。
        
        **【作業開始時（データが消えていた場合）】**
        1.  **「⚙️ データ管理」**タブを開きます。
        2.  「📥 データ復元」に、前回保存したファイルをアップロードします。
        """)

# --- メイン処理 ---
def main():
    init_session_state()
    st.title(PAGE_TITLE)
    
    # タブ構成
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📦 Tシャツ在庫", 
        "🏷️ タグ管理", 
        "📊 Tシャツ記録", 
        "⚙️ データ管理", 
        "📖 マニュアル"
    ])
    
    with tab1: inventory_tab()
    with tab2: tags_tab()
    with tab3: records_tab()
    with tab4: settings_tab()
    with tab5: manual_tab()

if __name__ == "__main__":
    main()
